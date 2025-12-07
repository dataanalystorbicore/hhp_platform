import streamlit as st
import pandas as pd
import gspread
from datetime import datetime
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font
from io import BytesIO


service_account_info = st.secrets["gcp_service_account"]
gc = gspread.service_account_from_dict(service_account_info)


Calendrier = gc.open("Official_Calendar")
Calendrier_worksheet = Calendrier.sheet1
Calendrier_values = Calendrier_worksheet.get_all_values()
df_Calendrier = pd.DataFrame(Calendrier_values[1:], columns=Calendrier_values[0])

# reads old visits report
visitreport = gc.open("visitsreport")
visitreport_worksheet = visitreport.sheet1


# reads database
database = gc.open("database")
database_worksheet = database.sheet1


def download_bbeinfo():
    bbe_info = gc.open("bbe_info")
    bbe_worksheet = bbe_info.sheet1
    bbe_values = bbe_worksheet.get_all_values()
    df_bbe = pd.DataFrame(bbe_values[1:], columns=bbe_values[0])

    return df_bbe


# sample rows for wilayas
def select_and_sample_rows_for_wilayas(df, BBE = None):

    if BBE != None:
        if (BBE in (df['BBE (ORB)'].unique())):
            df_filtered = df[df['BBE (ORB)'] == BBE]
        else:
            st.write("BBE didn't work yesterday")
    

    wilaya_rows = df.drop_duplicates(subset=['Wilaya'])
    target_rows = 50
    num_to_sample = target_rows - len(wilaya_rows)

    if num_to_sample <= 0:
        return wilaya_rows.reset_index(drop=True)

    remaining_rows = df[~df.index.isin(wilaya_rows.index)]

    if len(remaining_rows) < num_to_sample:
        sampled_rows = remaining_rows
    else:
        sampled_rows = remaining_rows.sample(n=num_to_sample, replace=False, random_state=1)

    if "df_filtered" in locals() or "df_filtered" in globals():

        final_df = pd.concat([wilaya_rows, sampled_rows, df_filtered]).reset_index(drop=True)
    else:
        final_df = pd.concat([wilaya_rows, sampled_rows]).reset_index(drop=True)

    final_df = final_df.sample(frac=1).reset_index(drop=True)

    return final_df



def cleaning(uploaded_file, BBE_CODE = None):

    if uploaded_file:
        try:
            st.write("File uploaded, processing...")

            df_bbe = download_bbeinfo()

            # --- Load visits
            data_visits = pd.read_excel(uploaded_file)

            data = database_worksheet.get_all_values()
            data_pos = pd.DataFrame(data[1:], columns=data[0])

            data_pos = data_pos[~((data_pos["PropriétairePhone"] == '0') & (data_pos["GérantPhone"] == '0' ))]

            if (data_visits['Username'] == 'test').any():
                data_visits = data_visits[data_visits['Username'] != 'test']

            if (data_visits['Closed'] == 'YES').any():
                data_visits = data_visits[data_visits['Closed'] != 'YES']

            data_visits = data_visits.merge(df_bbe[['BBE (ORB)', 'Username']], on = 'Username', how = 'left')

            data_visits["DISTRICT ID"] = (
                data_visits["Region"].astype(str) +
                data_visits["District"].astype(str).str.zfill(2) +
                data_visits["Territory"].astype(str) +
                data_visits["Code"].astype(str).str.zfill(5)
            ).astype("int64")

            if data_visits['Site ID'].isnull().any():
                lookup = dict(zip(data_pos['DISTRICT ID'].astype(int), data_pos['Site ID']))
                data_visits['Site ID'] = data_visits['Site ID'].fillna(data_visits['DISTRICT ID'].map(lookup))

            data_pos["Nom_complet_proprio"] = data_pos["PropriétaireLastname"] + "_" + data_pos["PropriétaireFirstname"]
            data_pos["Nom_complet_gerant"] = data_pos["GérantLastname"] + "_" + data_pos["GérantFirstname"]

            data_visits = data_visits.merge(
                data_pos[['Site ID','PropriétairePhone','GérantPhone',
                          'Nom_complet_proprio','Nom_complet_gerant',
                          'Pos Type']],
                on='Site ID', how='left'
            )

            data_visits = data_visits.drop_duplicates(subset=['Site ID'])
            data_visits['BBE (ORB)'] = data_visits['BBE (ORB)'].astype(str)
            data_visits = data_visits[['Region','BBE (ORB)','Wilaya',
                                       'Commune','Pos Type','Site ID','Name',
                                       'Nom_complet_proprio','PropriétairePhone',
                                       'Nom_complet_gerant','GérantPhone','Address']]
            

            Banned_stores = ['C070065974','C070066144','C000000408','C000000135']

            data_visits = data_visits[~data_visits['Site ID'].isin(Banned_stores)]

            for col in ['DATE','Merchandiser visit','Sales request for the week','POP S25',
                'Relationship with merchandiser','REMARK']:
                data_visits[col] = None     

            # Exclude already saved sites
            report_data = visitreport_worksheet.get_all_values()
            done_calls = pd.DataFrame(report_data[1:],columns=[
                'Region','BBE (ORB)','Wilaya','Commune','Pos Type',
                'Site ID','Name','Nom_complet_proprio','PropriétairePhone',
                'Nom_complet_gerant','GérantPhone','Address','DATE',
                'Merchandiser visit','Sales request for the week','POP S25',
                'Relationship with merchandiser','REMARK'])

            data_visits = data_visits.rename(columns = {"Address" : "POS Adress"})


            if not done_calls.empty:
                
                done_calls["DATE"] = pd.to_datetime(done_calls["DATE"])
                df_Calendrier["DATE"] = pd.to_datetime(df_Calendrier["DATE"])

                df_Calendrier["Week"] = pd.to_numeric(df_Calendrier["Week"], errors="coerce")

                done_calls = done_calls.merge(df_Calendrier, on="DATE", how="left")

                today = date.today()

                week_max_series = df_Calendrier.loc[df_Calendrier["DATE"] == pd.to_datetime(today), "Week"]

                if not week_max_series.empty:
                    week_max = week_max_series.iloc[0]
                    week_min = week_max - 3

                    st.write(f"Filtering weeks between {week_min} and {week_max}")

                    done_calls = done_calls[(done_calls["Week"] >= week_min) & (done_calls["Week"] <= week_max)]

                    saved_sites = done_calls['Site ID'].unique()
                    data_visits = data_visits[~data_visits['Site ID'].isin(saved_sites)]
                else:
                    st.warning(f"Date {today} not found in calendar.")

            data_visits = data_visits.dropna(axis = 0, subset = ["PropriétairePhone","GérantPhone"])

            data_visits = select_and_sample_rows_for_wilayas(data_visits,BBE_CODE)

            st.success("Rapport Des Appels Est Prét")

            # --- STEP 1: Write DataFrame to temp buffer
            tmp_buffer = BytesIO()
            with pd.ExcelWriter(tmp_buffer, engine="openpyxl") as writer:
                data_visits.to_excel(writer, index=False, sheet_name="Sheet1")
            tmp_buffer.seek(0)

            # --- STEP 2: Open with openpyxl
            wb = openpyxl.load_workbook(tmp_buffer)
            ws = wb.active

            green_fill = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")
            blue_fill = PatternFill(start_color="FF757171", end_color="FF757171", fill_type="solid")
            red_fill = PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid")
            white_bold_font = Font(color="FFFFFFFF", bold=True)
            font = Font(name='Calibri', size=12)

            for col_index, cell in enumerate(ws[1], 1):
                cell.font = white_bold_font
                if col_index <= 6:
                    cell.fill = green_fill
                elif 7 <= col_index <= 12:
                    cell.fill = blue_fill
                elif 13 <= col_index <= 20:
                    cell.fill = red_fill

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = font


            # --- STEP 3: Save to NEW buffer
            final_buffer = BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

            # --- STEP 4: Download
            st.download_button(
                label="📥 Download Call Report",
                data=final_buffer,
                file_name="Phone Calls.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Failed to process file: {e}")
