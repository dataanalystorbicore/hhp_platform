import streamlit as st
import pandas as pd
import gspread
import openpyxl
from openpyxl.styles import PatternFill, Font
from io import BytesIO


service_account_info = st.secrets["gcp_service_account"]
gc = gspread.service_account_from_dict(service_account_info)


visitreport = gc.open("visitsreport")
visitreport_worksheet = visitreport.sheet1


database = gc.open("database")
database_worksheet = database.sheet1

def select_and_sample_rows_for_wilayas(df):
   
    wilaya_rows = df.drop_duplicates(subset=['Wilaya'])
    
    target_rows = 55
    num_to_sample = target_rows - len(wilaya_rows) 

    if num_to_sample <= 0:
        return wilaya_rows.reset_index(drop=True)
    
    remaining_rows = df[~df.index.isin(wilaya_rows.index)]
    
    if len(remaining_rows) < num_to_sample:
        sampled_rows = remaining_rows
    else:
        sampled_rows = remaining_rows.sample(n=num_to_sample, replace=False, random_state=1)
    
    final_df = pd.concat([wilaya_rows, sampled_rows]).reset_index(drop=True)

    final_df = final_df.sample(frac=1).reset_index(drop=True)
    
    return final_df

# --- How to integrate this into your existin

def color_excel_header_in_memory(buffer):
    """
    Colors the header of an Excel file in a BytesIO buffer.
    """
    try:
        # Load the workbook from the buffer
        wb = openpyxl.load_workbook(buffer)
        ws = wb.active

        # Define the colors
        green_fill = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")
        blue_fill = PatternFill(start_color="FF757171", end_color="FF757171", fill_type="solid")
        red_fill = PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid")

        white_bold_font = Font(color="FFFFFFFF", bold=True)
        
        # Apply the colors to the header row
        header_row = ws[1]
        for col_index, cell in enumerate(header_row, 1):
            cell.font = white_bold_font
            if col_index <= 6:
                cell.fill = green_fill
            elif 7 <= col_index <= 12:
                cell.fill = blue_fill
            elif 13 <= col_index <= 20:
                cell.fill = red_fill
            else:
                pass

        # Save the workbook back to the buffer
        wb.save(buffer)
        buffer.seek(0)  # Rewind the buffer to the beginning
        return True
    except Exception as e:
        st.error(f"Failed to color the Excel header: {e}")
        return False


def cleaning(uploaded_file):

    if uploaded_file:
        try:

            st.write("file uploaded processing ...")
            
            samsung_orbicore = {
            "R1ORB002": "ORB002#R#44",
            "R1ORB009": "ORB009#R#16",
            "R1ORB004": "ORB004#R#16",
            "R1ORB032": "ORB032#R#42",
            "R1ORB007": "ORB007#R#16",
            "R1ORB016": "ORB016#R#35",
            "R1ORB043": "ORB043#WS#16",
            "R1ORB036": "ORB036#WS#16",
            "R1ORB003": "ORB003#R#16",
            "R2ORB015": "ORB015#R#34",
            "R3ORB022": "ORB022#R#29",
            "R3ORB034": "ORB034#R#13",
            "R2ORB039": "ORB039#R#23",
            "R2ORB011": "ORB011#R#05",
            "R2ORB060": "ORB060#R#05",
            "R1ORB020": "ORB020#R#17",
            "R2ORB013": "ORB013#R#06",
            "R1ORB042": "ORB042#R#09",
            "R2ORB037": "ORB037#WS#19",
            "R3ORB017": "ORB017#R#02",
            "R2ORB018": "ORB018#R#25",
            "R2ORB019": "ORB019#R#25",
            "R2ORB030": "ORB030#R#12",
            "R3ORB025": "ORB025#R#31",
            "R3ORB024": "ORB024#R#27",
            "R3ORB038": "ORB038#WS#31",
            "R2ORB052": "ORB052#R#SUP05",
            "R1ORB041": "ORB041#R#15",
            "R2ORB061": "ORB060#R#05"
            }

            
            data_visits = pd.read_excel(uploaded_file)
            
            
            data = database_worksheet.get_all_values()
            data_pos = pd.DataFrame(data[1:], columns=data[0])
            

            if (data_visits['Username'] == 'test').any():
                data_visits = data_visits[data_visits['Username'] != 'test']

            if (data_visits['Closed'] == 'YES').any():
                data_visits = data_visits[data_visits['Closed'] != 'YES']

                
            data_visits['BBE_ID'] = data_visits['Username'].apply(lambda x: samsung_orbicore.get(x, None))
            data_visits["Pos id"] = (
                data_visits["Region"].astype(str) +
                data_visits["District"].astype(str).str.zfill(2) +
                data_visits["Territory"].astype(str) +
                data_visits["Code"].astype(str).str.zfill(5)
            ).astype(int)

            if data_visits['Site ID'].isnull().any():

                lookup = dict(zip(data_pos['Pos id'].astype(int), data_pos['Site ID']))
                # Map Site ID from lookup only where missing
                data_visits['Site ID'] = data_visits['Site ID'].fillna(data_visits['Pos id'].map(lookup))
    
            data_visits = data_visits.merge(data_pos[['Site ID','num_proprio','num_gerant','Nom_complet_proprio','Nom_complet_gerant','Pos Type','Wilaya(48)']], 
                                            on='Site ID', how='left')
            data_visits = data_visits.drop_duplicates(subset = ['Site ID'])
            data_visits = data_visits[['Region','BBE_ID','Wilaya','Wilaya(48)','District','Pos Type','Site ID','Name','Nom_complet_proprio','num_proprio',
                                       'Nom_complet_gerant', 'num_gerant','Address']] 
            new_columns = ['Date','Visit','Sales','POP','Relationship','Remark']
            for col in new_columns:
                data_visits[col] = None     
                        
            report_data = visitreport_worksheet.get_all_values()
            done_calls = pd.DataFrame(report_data[1:],columns = ['Region','BBE_ID','Wilaya','Wilaya(48)','District','Pos Type','Site ID','Name','Nom_complet_proprio',
                                        'num_proprio','Nom_complet_gerant', 'num_gerant','Address','Date','Visit','Sales','POP','Relationship','Remark'])

            if not done_calls.empty:
                saved_sites = done_calls['Site ID'].unique()
                data_visits = data_visits[~data_visits['Site ID'].isin(saved_sites)]

            data_visits = data_visits.fillna(" ")
            #data_visits = select_and_sample_rows_for_wilayas(data_visits)

           
            st.success("New Cleaned visits were Added succesfully")

            # --- New Logic: Save to a BytesIO buffer ---
            output_buffer = BytesIO()
            data_visits.to_excel(output_buffer, index=False, engine='openpyxl')
            
            # Call the coloring function on the buffer
            coloring_success = color_excel_header_in_memory(output_buffer)
            
            if coloring_success:
                
                st.download_button(
                    label="Download Call Report",
                    data=output_buffer,
                    file_name="cleaned_visits.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            

        except Exception as e:
            st.error(f"Failed to read file: {e}")


       